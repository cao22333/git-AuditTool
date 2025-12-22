import pandas as pd
import os
import xlwings as xw
from openpyxl import load_workbook # 用于读取映射表

class ReconciliationManager:
    """
    科目勾稽管理类。
    封装了薪酬勾稽和资产折旧摊销勾稽的核心逻辑，不包含UI相关代码。
    """
    def __init__(self, logger=None):
        self.logger = logger if logger else self._default_logger # 注入日志器
        self.xw_app = None # xlwings 应用程序实例

    def _default_logger(self, message, level="INFO"):
        """默认日志器，如果未提供外部日志器则打印到控制台"""
        print(f"[{level}] {message}")

    def _init_xlwings_app(self):
        """初始化 xlwings 应用程序实例"""
        if self.xw_app is None:
            self.xw_app = xw.App(visible=False, add_book=False)
            self.xw_app.display_alerts = False 
            self.xw_app.screen_updating = False
            self.logger("xlwings 应用程序已初始化。", "INFO")

    def _quit_xlwings_app(self):
        """退出 xlwings 应用程序实例"""
        if self.xw_app is not None:
            self.xw_app.quit()
            self.xw_app = None
            self.logger("xlwings 应用程序已退出。", "INFO")

    def payroll_reconciliation(self, zt_path, mapping_file, output_base_dir, progress_callback=None):
        """
        执行薪酬勾稽。
        zt_path: 账套文件所在目录。
        mapping_file: 勾稽映射表Excel文件路径。
        output_base_dir: 输出底稿文件存放的目录。
        """
        self.logger("开始执行薪酬勾稽...", "INFO")
        self._init_xlwings_app() # 确保 xlwings 应用已启动

        try:
            # 读取映射表 Sheet5
            wb_mapping = load_workbook(mapping_file, data_only=True)
            ws_mapping = wb_mapping['Sheet5']

            pathdict = {}  # 存储账套文件路径和对应的输出路径
            for i in range(2, ws_mapping.max_row + 1): # 遍历所有行
                if ws_mapping[f'B{i}'].value and ws_mapping[f'D{i}'].value: # 确保单元格有值
                    pathi1 = os.path.join(zt_path, ws_mapping[f'B{i}'].value)
                    pathi2 = os.path.join(output_base_dir, ws_mapping[f'D{i}'].value)
                    pathdict[pathi1] = pathi2
            
            if not pathdict:
                return False, "映射表中未找到有效的账套文件路径和输出路径。"

            zcfy_dict = {
                '应付职工薪酬-业务维护提成超量奖金': '业务维护提成超量奖金',
                '应付职工薪酬-超任务奖金': '超任务奖',
                '应付职工薪酬-其他专项奖': '其他专项奖',
                '应付职工薪酬-工资': '工资',
                '应付职工薪酬-福利': '福利费',
                '经费': '工会经费',
                '保险': '社会保险',
                '应付职工薪酬-住房公积金': '住房公积金',
                '应付职工薪酬-辞退福利': '辞退福利',
                '应付职工薪酬-预留绩效薪金': '预留绩效薪金'
            }

            zcfy_dict_1 = { # Excel 写入的行号映射
                '应付职工薪酬-业务维护提成超量奖金': 7,
                '应付职工薪酬-超任务奖金': 6,
                '应付职工薪酬-其他专项奖': 9,
                '应付职工薪酬-工资': 5,
                '应付职工薪酬-福利': 10,
                '经费': 19,
                '保险': 11,
                '应付职工薪酬-住房公积金': 18,
                '应付职工薪酬-辞退福利': 22,
                '应付职工薪酬-预留绩效薪金': 8
            }

            total_tasks = len(pathdict)
            for idx, (xsz_file_path, output_excel_path) in enumerate(pathdict.items()):
                if progress_callback:
                    progress_callback(int((idx / total_tasks) * 100), f"正在处理 {os.path.basename(xsz_file_path)}...")

                self.logger(f"正在处理账套文件: {xsz_file_path}", "INFO")
                if not os.path.exists(xsz_file_path):
                    self.logger(f"警告: 账套文件不存在，跳过: {xsz_file_path}", "WARNING")
                    continue
                
                # 读取序时账
                xsz_df = pd.read_excel(xsz_file_path, dtype={'科目编码': str})
                
                # 确保 '借方' 和 '贷方' 列为数值类型
                for col in ['借方', '贷方']:
                    if col in xsz_df.columns:
                        xsz_df[col] = pd.to_numeric(
                            xsz_df[col].astype(str).str.replace(',', ''),
                            errors='coerce'
                        ).fillna(0)
                    else:
                        self.logger(f"警告: 账套文件 '{os.path.basename(xsz_file_path)}' 中缺少列 '{col}'。", "WARNING")
                        xsz_df[col] = 0 # 填充默认值
                
                # 打开输出Excel文件
                if not os.path.exists(output_excel_path):
                    self.logger(f"错误: 输出底稿文件不存在，无法写入: {output_excel_path}", "ERROR")
                    continue
                
                wb_ok = self.xw_app.books.open(output_excel_path)
                
                for key_1, value_1 in zcfy_dict.items():
                    # 检查科目名称中是否存在关键字
                    if '科目名称' in xsz_df.columns:
                        mask = xsz_df['科目名称'].astype(str).str.contains(key_1, na=False)
                        if mask.any():
                            row_value = zcfy_dict_1.get(key_1)
                            if row_value:
                                # 统计薪酬贷方发生额
                                zc_ljzj_df = xsz_df.loc[mask & (xsz_df['科目编码'].astype(str).str.startswith('2211', na=False))]
                                total_amount = zc_ljzj_df['贷方'].sum()
                                wb_ok.sheets['应付职工薪酬分配检查情况表'].range(f'I{row_value}').value = total_amount
                                self.logger(f"已更新 '{key_1}' 的贷方发生额到 {output_excel_path}。", "INFO")
                            else:
                                self.logger(f"警告: 未找到科目 '{key_1}' 对应的写入行号。", "WARNING")
                        else:
                            self.logger(f"'{os.path.basename(xsz_file_path)}' 中不存在包含【{key_1}】的数据。", "INFO")
                    else:
                        self.logger(f"警告: 账套文件 '{os.path.basename(xsz_file_path)}' 中缺少列 '科目名称'。", "WARNING")

                wb_ok.save()
                wb_ok.close()
                self.logger(f"'{os.path.basename(xsz_file_path)}' 勾稽表填写完成。", "SUCCESS")
            
            return True, "薪酬勾稽执行完成！"
        except Exception as e:
            self.logger(f"薪酬勾稽失败: {str(e)}", "ERROR")
            return False, f"薪酬勾稽失败: {str(e)}"
        finally:
            self._quit_xlwings_app() # 确保 xlwings 应用被关闭

    def asset_depreciation_reconciliation(self, yb_path, zjgjb_mb, output_base_dir, progress_callback=None):
        """
        执行资产折旧摊销勾稽。
        yb_path: 原始报表文件所在目录。
        zjgjb_mb: 折旧分配表模板文件路径。
        output_base_dir: 输出底稿文件存放的目录。
        """
        self.logger("开始执行资产折旧摊销勾稽...", "INFO")
        self._init_xlwings_app() # 确保 xlwings 应用已启动

        try:
            yb_filelist = [f for f in os.listdir(yb_path) if f.endswith('.xlsx') or f.endswith('.xls')]
            if not yb_filelist:
                return False, f"原始报表目录 '{yb_path}' 中未找到Excel文件。"

            zcfy_dict = {
                '固定资产折旧费': '1602',
                '使用权资产折旧': '1803',
                '无形资产摊销': '1702',
                '长期待摊费用摊销': '1801'
            }

            total_tasks = len(yb_filelist)
            for idx, file_name in enumerate(yb_filelist):
                if progress_callback:
                    progress_callback(int((idx / total_tasks) * 100), f"正在处理 {file_name}...")

                file_path = os.path.join(yb_path, file_name)
                self.logger(f"正在处理报表文件: {file_path}", "INFO")

                # 读取报表文件中的费用明细表
                try:
                    yb_df_xsfy = pd.read_excel(file_path, sheet_name='C03_销售费用明细表')
                    yb_df_glfy = pd.read_excel(file_path, sheet_name='C04_管理费用明细表')
                    yb_df_yffy = pd.read_excel(file_path, sheet_name='C05_研发费用明细表')
                    yb_df_zzfy = pd.read_excel(file_path, sheet_name='C07_制造费用明细表')
                    yb_df_sccb = pd.read_excel(file_path, sheet_name='C08_生产成本明细表')
                except Exception as e:
                    self.logger(f"读取报表文件 '{file_name}' 失败: {str(e)}，跳过。", "WARNING")
                    continue
                
                # 获取编制单位和公司名称
                yb_bzdw = yb_df_xsfy.iloc[15, 4]
                yb_gsmc = yb_bzdw[5:]
                self.logger(f"公司名称: {yb_gsmc}", "INFO")

                # 提取费用明细数据
                yb_df_xsfy_sxjg = yb_df_xsfy.iloc[17:98, [4, 5, 6, 7, 8, 9, 10, 11, 12, 13]].fillna(0)
                yb_df_glfy_sxjg = yb_df_glfy.iloc[22:103, [4, 5, 6, 7, 8, 9, 10, 11, 12, 13]].fillna(0)
                yb_df_yffy_sxjg = yb_df_yffy.iloc[22:103, [4, 5, 6, 7, 8, 9, 10, 11, 12, 13]].fillna(0)
                yb_df_zzfy_sxjg = yb_df_zzfy.iloc[22:103, [4, 5, 6, 7, 8, 9, 10, 11, 12, 13]].fillna(0)
                yb_df_sccb_sxjg = yb_df_sccb.iloc[22:103, [4, 5, 6, 7, 8, 9, 10, 11, 12, 13]].fillna(0)

                # 打开折旧分配表模板
                if not os.path.exists(zjgjb_mb):
                    self.logger(f"错误: 折旧分配表模板不存在: {zjgjb_mb}", "ERROR")
                    return False, f"折旧分配表模板不存在: {zjgjb_mb}"

                wb_ok = self.xw_app.books.open(zjgjb_mb)
                
                for key, value in zcfy_dict.items():
                    yf_df_xsfy_jg = yb_df_xsfy_sxjg[yb_df_xsfy_sxjg.iloc[:, 0] == key].T
                    yf_df_glfy_jg = yb_df_glfy_sxjg[yb_df_glfy_sxjg.iloc[:, 0] == key].T
                    yf_df_yffy_jg = yb_df_yffy_sxjg[yb_df_yffy_sxjg.iloc[:, 0] == key].T
                    yf_df_zzfy_jg = yb_df_zzfy_sxjg[yb_df_zzfy_sxjg.iloc[:, 0] == key].T
                    yf_df_sccb_jg = yb_df_sccb_sxjg[yb_df_sccb_sxjg.iloc[:, 0] == key].T
                    
                    # 复制名为“折旧分配测算表”的工作表
                    sheet_to_copy = wb_ok.sheets['折旧分配测算表']
                    new_sheet = sheet_to_copy.copy()
                    new_sheet.name = f'{key}_勾稽表'
                    
                    # 将 DataFrame 写入工作表的指定单元格
                    if not yf_df_glfy_jg.empty:
                        wb_ok.sheets[f'{key}_勾稽表'].range('C6').options(index=False, header=False).value = yf_df_glfy_jg.iloc[1:]
                    if not yf_df_xsfy_jg.empty:
                        wb_ok.sheets[f'{key}_勾稽表'].range('D6').options(index=False, header=False).value = yf_df_xsfy_jg.iloc[1:]
                    if not yf_df_yffy_jg.empty:
                        wb_ok.sheets[f'{key}_勾稽表'].range('E6').options(index=False, header=False).value = yf_df_yffy_jg.iloc[1:]
                    if not yf_df_zzfy_jg.empty:
                        wb_ok.sheets[f'{key}_勾稽表'].range('F6').options(index=False, header=False).value = yf_df_zzfy_jg.iloc[1:]
                    if not yf_df_sccb_jg.empty:
                        wb_ok.sheets[f'{key}_勾稽表'].range('G6').options(index=False, header=False).value = yf_df_sccb_jg.iloc[1:]
                    
                    self.logger(f"已更新 '{key}' 的勾稽数据到 '{key}_勾稽表'。", "INFO")

                output_file = os.path.join(output_base_dir, f'{yb_gsmc}_资产勾稽核对表.xlsx')
                wb_ok.save(output_file)
                wb_ok.close()
                self.logger(f"'{yb_gsmc}' 资产勾稽表填写完成，保存至: {output_file}", "SUCCESS")
            
            return True, "资产折旧摊销勾稽执行完成！"
        except Exception as e:
            self.logger(f"资产折旧摊销勾稽失败: {str(e)}", "ERROR")
            return False, f"资产折旧摊销勾稽失败: {str(e)}"
        finally:
            self._quit_xlwings_app() # 确保 xlwings 应用被关闭